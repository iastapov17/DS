from openpyxl import load_workbook
import json

def split_name(fio, info):
    fio = fio.replace("'", '').replace('X', 'KS').replace("OXA","OKSA").replace("TC", "TS").replace("EX", "EKS").replace("AY", "AI").replace("IY", "II").replace("TCOV", "TSOV").replace("YA","IA").replace("YU", "IU").replace("EY", "EI").split(' ')
    if len(fio) == 2:
        info["Name"] = fio[0]
        info["Surname"] = fio[1]
    elif len(fio) == 3:
        if len(fio[0]) == 1:
            info["Secname"] = fio[0].replace('Y', 'I') + '.'
            info["Name"] = fio[1]
            info["Surname"] = fio[2]
        elif len(fio[1]) == 1:
            info["Secname"] = fio[1].replace('Y', 'I') + '.'
            info["Name"] = fio[0]
            info["Surname"] = fio[2]
        else:
            info["Secname"] = fio[2].replace('Y', 'I') + '.'
            info["Name"] = fio[0]
            info["Surname"] = fio[1]

def get_flight_ticket(file):
    fields = {
        "FIO": (3, 2),
        "Sex": (3, 1),
        "Sequence": (1, 8),
        "Sequence_word": (3, 8),
        "Flight": (5, 1),
        "From": (5, 4),
        "To": (5, 8),
        "From_red": (7, 4),
        "To_red": (7, 8),
        "Gate": (7, 2),
        "Date": (9, 1),
        "Time": (9, 3),
        "Seat": (11, 8),
        "PNR": (13, 2),
        "Ticket": (13, 5)
    }

    info_in_one_file = []
    wb = load_workbook("/home/roman/Загрузки/Airlines/passdot/" + file)
    for sheet_name in wb.get_sheet_names():
        people_info = {}
        for field in fields:
            if field == "FIO":
                split_name(wb[sheet_name].cell(fields[field][0], fields[field][1]).value, people_info)
                continue
            if field == "Sex":
                sex = wb[sheet_name].cell(fields[field][0], fields[field][1]).value
                sex = 1 if sex == "MRS" else 0
                people_info["Sex"] = sex
                continue
            people_info[field] = wb[sheet_name].cell(fields[field][0], fields[field][1]).value
        info_in_one_file.append(people_info)
    wb.close()
    return info_in_one_file


def get_ticket_json():
    mypath = "/home/roman/Загрузки/Airlines/passdot"
    files = [f for f in listdir(mypath) if isfile(join(mypath, f))]

    all_data = []

    for file in files:
        with open(file, 'r') as f:
            data = json.load(f)
        all_data.append(data)

    return all_data
